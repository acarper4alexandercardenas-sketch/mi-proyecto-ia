import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def leer_equipo(archivo):
    equipo = []
    with open(archivo, encoding='utf-8') as f:
        lector = csv.DictReader(f)
        for fila in lector:
            equipo.append(fila)
            
    return equipo

def mostrar_resumen(equipo):
    total_personas = 0
    con_coordinador = 0
    sin_coordinador = 0

    print("=" * 45)
    print("RESUMEN DEL EQUIPO DE TI")
    print("=" * 45)

    for area in equipo:
        personas = int(area['personas'])
        total_personas += personas
        if area['coordinador'] == 'Si':
            con_coordinador += 1
        else:
            sin_coordinador += 1
        print(f"{area['area']:20} | {personas} personas | {area['sistema']}")

    print("=" * 45)
    print(f"Total personas     : {total_personas}")
    print(f"Areas con coord.   : {con_coordinador}")
    print(f"Areas sin coord.   : {sin_coordinador}")
    print("=" * 45)



def filtrar_por_sistema(equipo_filtrado, sistema):
    print(f"\nAreas que usan {sistema}:")
    print("-" * 35)
    for area in equipo_filtrado:
        print(f"- {area['area']} ({area['personas']} personas)")
        

def exportar_excel(equipo, archivo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipo TI"

    # Encabezados
    encabezados = ["Area", "Personas", "Sistema", "Coordinador"]
    for col, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        celda.alignment = Alignment(horizontal="center")

    # Datos
    for row, area in enumerate(equipo, 2):
        ws.cell(row=row, column=1, value=area['area'])
        ws.cell(row=row, column=2, value=int(area['personas']))
        ws.cell(row=row, column=3, value=area['sistema'])
        ws.cell(row=row, column=4, value=area['coordinador'])

    # Total
    total = sum(int(a['personas']) for a in equipo)
    fila_total = len(equipo) + 2
    ws.cell(row=fila_total, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=fila_total, column=2, value=total).font = Font(bold=True)

    # Ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

    wb.save(archivo)
    print(f"\nArchivo Excel exportado: {archivo}")



equipo = leer_equipo('equipo.csv')
mostrar_resumen(equipo)


sistema_buscar = input("\n¿Qué sistema quieres buscar? ")

# Filtramos UNA sola vez
equipo_filtrado = [a for a in equipo if sistema_buscar.upper() in a['sistema'].upper()]

# Usamos el mismo resultado para ambas cosas
filtrar_por_sistema(equipo_filtrado, sistema_buscar)

nombre_archivo = f"reporte_{sistema_buscar.lower()}.xlsx"
exportar_excel(equipo_filtrado, nombre_archivo)


